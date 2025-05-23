import random
import string, os, csv
import openpyxl
from django.conf import settings
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Sum, Count
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.core.mail import EmailMessage, send_mail
from weasyprint import HTML
from .models import Employee, Payslip, Salary
from .forms import EmployeeForm, SalaryForm, EmployeeUploadForm,\
    PayrollUploadForm, MonthlyPayrollSummaryForm,\
        EmployeeCompensationReportForm


# Create your views here.
@login_required
def dashboard(request):
    total_employees = Employee.objects.count()
    total_salary_last_month = Payslip.objects.filter(date__month=timezone.now().month-1).aggregate(Sum('net_pay'))['net_pay__sum'] or 0
    payrolls_processed = Payslip.objects.count()

    # Total salary paid per month
    salary_per_month = Payslip.objects.values('date__month').annotate(total_salary=Sum('net_pay')).order_by('date__month')

    # Total payrolls processed per month
    payrolls_per_month = Payslip.objects.values('date__month').annotate(total_payrolls=Count('id')).order_by('date__month')

    # Fetch list of employees and payslips
    employees = Employee.objects.all()
    payslips = Payslip.objects.all()

    context = {
        'total_employees': total_employees,
        'total_salary_last_month': total_salary_last_month,
        'payrolls_processed': payrolls_processed,
        'salary_per_month': salary_per_month,
        'payrolls_per_month': payrolls_per_month,
        'employees': employees,
        'payslips': payslips,
    }

    return render(request, 'core/dashboard.html', context)


@login_required
def management_dashboard(request):
    return render(request, 'core/management_dashboard.html')


@login_required
def employee_list(request):
    employees = Employee.objects.all()
    paginator = Paginator(employees, 10)  # Show 10 employees per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    form = EmployeeForm()
    return render(request, 'core/employee_dashboard.html', {'page_obj': page_obj, 'form': form})

@login_required
def employee_add(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            phone_number = form.cleaned_data['phone_number']
            if not Employee.objects.filter(name=name, email=email, phone_number=phone_number).exists():
                employee = form.save(commit=False)
                employee.employee_code = generate_reference_code(employee)
                employee.save()
                messages.success(request, 'Employee added successfully.')
            return redirect('core:employee_list')
    return redirect('core:employee_list')

@login_required
def employee_update(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            return redirect('core:employee_list')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'core/edit_employee.html', {'form': form, 'employee': employee})

@login_required
def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        return redirect('core:employee_list')
    return redirect('core:employee_list')

def generate_reference_code(employee):
    count = Employee.objects.filter(location=employee.location, staff_type=employee.staff_type).count() + 1
    random_code = ''.join(random.choices(string.ascii_letters + string.digits, k=3))
    return f"RUN/MP/{employee.staff_type}/{random_code}/{count}"

@login_required
def salary_list(request):
    salaries = Salary.objects.all()
    paginator = Paginator(salaries, 10)  # Show 10 salaries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    form = SalaryForm()
    return render(request, 'core/salary_dashboard.html', {'page_obj': page_obj, 'form': form})

@login_required
def salary_add(request):
    if request.method == 'POST':
        form = SalaryForm(request.POST)
        if form.is_valid():
            salary = form.save()
            # Calculate gross pay, total deductions, and net pay
            gross_pay = (
                salary.basic_pay +
                salary.rent_allowance +
                salary.transport_allowance +
                salary.meal_allowance +
                salary.utility_allowance +
                salary.other_allowances +
                salary.overtime +
                salary.arrears
            )
            total_deductions = (
                salary.staff_loan_deduction +
                salary.rent_deduction +
                salary.medical_deduction +
                salary.pension_contribution +
                salary.shortage_deduction +
                salary.development_levy +
                salary.personal_income_tax
            )
            net_pay = gross_pay - total_deductions

            # Create a corresponding payslip record
            Payslip.objects.create(
                employee=salary.employee,
                salary=salary,
                date=salary.salary_date,
                gross_pay=gross_pay,
                total_deductions=total_deductions,
                net_pay=net_pay
            )
            return redirect('core:salary_list')
    return redirect('core:salary_list')
    

@login_required
def salary_update(request, pk):
    salary = get_object_or_404(Salary, pk=pk)
    if request.method == 'POST':
        form = SalaryForm(request.POST, instance=salary)
        if form.is_valid():
            form.save()
            return redirect('core:salary_list')
    else:
        form = SalaryForm(instance=salary)
    return render(request, 'core/edit_salary.html', {'form': form, 'salary': salary})

@login_required
def salary_delete(request, pk):
    salary = get_object_or_404(Salary, pk=pk)
    if request.method == 'POST':
        salary.delete()
        return redirect('core:salary_list')
    return redirect('core:salary_list')


def generate_payslip_pdf(payslip):
    salary = payslip.salary
    # Calculate gross pay, total deductions, and net pay
    gross_pay = (
        salary.basic_pay +
        salary.rent_allowance +
        salary.transport_allowance +
        salary.meal_allowance +
        salary.utility_allowance +
        salary.other_allowances +
        salary.overtime +
        salary.arrears
    )
    total_deductions = (
        salary.staff_loan_deduction +
        salary.rent_deduction +
        salary.medical_deduction +
        salary.pension_contribution +
        salary.shortage_deduction +
        salary.development_levy +
        salary.personal_income_tax
    )
    net_pay = gross_pay - total_deductions

    context = {
        'payslip': payslip,
        'salary': salary,
        'gross_pay': gross_pay,
        'company_logo': os.path.join(settings.STATIC_ROOT, 'assets/img/brand/logo.png'),
        'employee': payslip.employee,
        'company_name': "Redeemer's University Cafetaria",
        'total_deductions': total_deductions,
        'net_pay': net_pay
    }
    html_string = render_to_string('core/payslip.html', context)
    html = HTML(string=html_string, base_url=settings.BASE_DIR)
    return html


@login_required
def generate_payslip(request, pk):
    payslip = get_object_or_404(Payslip, pk=pk)
    html = generate_payslip_pdf(payslip)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=payslip_{payslip.employee.name}_{payslip.date}.pdf'
    html.write_pdf(response)
    return response


@login_required
def payslip_list(request):
    payslips = Payslip.objects.all()
    paginator = Paginator(payslips, 10)  # Show 10 payslips per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'core/payslip_dashboard.html', {'page_obj': page_obj})

@login_required
def view_payslip(request, pk):
    payslip = get_object_or_404(Payslip, pk=pk)
    html = generate_payslip_pdf(payslip)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=payslip_{payslip.employee.name}_{payslip.date}.pdf'
    html.write_pdf(response)
    return response



@login_required
def send_payslip_email(request, pk):
    payslip = get_object_or_404(Payslip, pk=pk)
    html = generate_payslip_pdf(payslip)
    pdf = html.write_pdf()

    email = EmailMessage(
        subject=f'Payslip for {payslip.employee.name}',
        body=f"Dear {payslip.employee.name},\nPlease find attached your payslip for the month of {payslip.date.strftime('%B')}.",
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[payslip.employee.email],
    )
    email.attach(f'payslip_{payslip.employee.employee_code}_{payslip.date}.pdf', pdf, 'application/pdf')
    email.send()

    return redirect('core:payslip_list')


def test_email(request):
    subject = 'Test Email'
    message = 'This is a test email.'
    from_email = settings.DEFAULT_FROM_EMAIL
    recipient_list = ['info.jaybeloved@gmail.com']  # Replace with your email address

    try:
        send_mail(subject, message, from_email, recipient_list)
        return HttpResponse('Email sent successfully.')
    except Exception as e:
        return HttpResponse(f'Error sending email: {e}')


@login_required
def upload_employees(request):
    if request.method == 'POST':
        form = EmployeeUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['file']
            
            # Check file extension
            if not csv_file.name.endswith('.csv'):
                messages.error(request, 'Please upload a CSV file. The employee template requires a CSV format.')
                return render(request, 'core/upload_employees.html', {'form': form})
                
            try:
                decoded_file = csv_file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                
                # Check if the file is empty
                if not reader.fieldnames:
                    messages.error(request, 'The uploaded CSV file is empty or improperly formatted.')
                    return render(request, 'core/upload_employees.html', {'form': form})
                    
                # Verify required columns exist
                required_columns = ['name', 'email', 'staff_type', 'phone_number']
                missing_columns = [col for col in required_columns if col not in reader.fieldnames]
                if missing_columns:
                    messages.error(request, f'The CSV file is missing required columns: {", ".join(missing_columns)}')
                    return render(request, 'core/upload_employees.html', {'form': form})
                
                # Process rows
                success_count = 0
                error_count = 0
                errors = []
                
                for row_num, row in enumerate(reader, start=2):  # Start at 2 to account for header row
                    # Check required fields are not empty
                    empty_fields = [field for field in required_columns if not row.get(field, '').strip()]
                    
                    if empty_fields:
                        errors.append(f'Row {row_num}: Missing values for {", ".join(empty_fields)}')
                        error_count += 1
                        continue
                    
                    # Validate staff type
                    valid_staff_types = ['SS', 'CoS', 'CaS']
                    if row['staff_type'] not in valid_staff_types:
                        errors.append(f'Row {row_num}: Invalid staff_type "{row["staff_type"]}". Must be one of {", ".join(valid_staff_types)}')
                        error_count += 1
                        continue
                        
                    # Check for duplicate
                    if Employee.objects.filter(name=row['name'], email=row['email']).exists():
                        errors.append(f'Row {row_num}: Employee with name "{row["name"]}" and email "{row["email"]}" already exists')
                        error_count += 1
                        continue
                    
                    try:
                        # Create employee
                        employee = Employee.objects.create(
                            name=row['name'],
                            email=row['email'],
                            location=row.get('location', ''),
                            staff_type=row['staff_type'],
                            designation=row.get('designation', ''),
                            grade_level=row.get('grade_level', ''),
                            account_number=row.get('account_number', ''),
                            bank_name=row.get('bank_name', ''),
                            account_name=row.get('account_name', ''),
                            phone_number=row['phone_number'],
                        )
                        employee.employee_code = generate_reference_code(employee)
                        employee.save()
                        success_count += 1
                    except Exception as e:
                        errors.append(f'Row {row_num}: Error creating employee - {str(e)}')
                        error_count += 1
                
                # Report results
                if success_count > 0:
                    messages.success(request, f'Successfully imported {success_count} employee(s).')
                
                if error_count > 0:
                    error_message = f'Failed to import {error_count} employee(s). First 5 errors:'
                    for i, error in enumerate(errors[:5], 1):
                        error_message += f'<br>{i}. {error}'
                    if len(errors) > 5:
                        error_message += f'<br>...and {len(errors) - 5} more errors.'
                    messages.error(request, error_message)
                
                return redirect('core:employee_list')
                
            except UnicodeDecodeError:
                messages.error(request, 'Unable to read the file. Please ensure it is a valid CSV file with UTF-8 encoding.')
            except Exception as e:
                messages.error(request, f'An error occurred while processing the file: {str(e)}')
        else:
            messages.error(request, 'Failed to upload employees. Please check the file format.')
    else:
        form = EmployeeUploadForm()
    return render(request, 'core/upload_employees.html', {'form': form})

    
    
@login_required
def export_employee_template(request):
    # Define the headers for the CSV file
    headers = [
        'name', 'email', 'location', 'staff_type', 'designation', 
        'grade_level', 'account_number', 'bank_name', 'account_name', 'phone_number'
    ]

    # Create a CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=employee_template.csv'

    writer = csv.writer(response)
    writer.writerow(headers)

    return response

@login_required
def export_template(request):
    employees = Employee.objects.all()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Payroll Template'

    # Add headers
    headers = [
        'employee_code', 'employee_name', 'basic_pay', 'rent_allowance', 'transport_allowance', 
        'meal_allowance', 'utility_allowance', 'other_allowances', 'overtime', 'arrears', 
        'staff_loan_deduction', 'rent_deduction', 'medical_deduction', 'pension_contribution', 
        'shortage_deduction', 'development_levy', 'personal_income_tax', 'no_of_days', 'actual_attendance'
    ]
    sheet.append(headers)

    # Add employee data
    for employee in employees:
        sheet.append([
            employee.employee_code, employee.name, '', '', '', '', '', '', '', '', 
            '', '', '', '', '', '', '','',''
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=payroll_template.xlsx'
    workbook.save(response)
    return response


@login_required
def import_payroll(request):
    if request.method == 'POST':
        form = PayrollUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            
            # Check file extension
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Please upload an Excel file (.xlsx or .xls). The payroll template requires Excel format.')
                return render(request, 'core/import_payroll.html', {'form': form})
                
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active
                
                # Check if the sheet has content
                if sheet.max_row <= 1:  # Only has header row or is empty
                    messages.error(request, 'The uploaded Excel file is empty or contains only headers.')
                    return render(request, 'core/import_payroll.html', {'form': form})
                
                # Verify required columns exist (check headers)
                header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                if len(header_row) < 2 or 'employee_code' not in header_row and 'employee_name' not in header_row:
                    messages.error(request, 'The Excel file is missing required headers. Please use the template provided by the system.')
                    return render(request, 'core/import_payroll.html', {'form': form})
                
                # Process rows
                success_count = 0
                error_count = 0
                errors = []
                
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    # Check for empty rows
                    if all(cell is None or cell == '' for cell in row):
                        continue
                        
                    if len(row) < 2 or not row[0]:  # Missing employee code
                        errors.append(f'Row {row_num}: Employee code is missing')
                        error_count += 1
                        continue
                        
                    employee_code = row[0]
                    try:
                        # Find employee
                        employee = Employee.objects.get(employee_code=employee_code)
                    except Employee.DoesNotExist:
                        errors.append(f'Row {row_num}: Employee with code "{employee_code}" does not exist')
                        error_count += 1
                        continue
                    
                    try:
                        # Safely get values and convert to appropriate types
                        basic_pay = float(row[2] or 0)
                        rent_allowance = float(row[3] or 0)
                        transport_allowance = float(row[4] or 0)
                        meal_allowance = float(row[5] or 0)
                        utility_allowance = float(row[6] or 0)
                        other_allowances = float(row[7] or 0)
                        overtime = float(row[8] or 0)
                        arrears = float(row[9] or 0)
                        staff_loan_deduction = float(row[10] or 0)
                        rent_deduction = float(row[11] or 0)
                        medical_deduction = float(row[12] or 0)
                        pension_contribution = float(row[13] or 0)
                        shortage_deduction = float(row[14] or 0)
                        development_levy = float(row[15] or 0)
                        personal_income_tax = float(row[16] or 0)
                        no_of_days = int(row[17] or 0)
                        actual_attendance = int(row[18] or 0)
                        
                        # Create salary record
                        salary = Salary.objects.create(
                            employee=employee,
                            basic_pay=basic_pay,
                            rent_allowance=rent_allowance,
                            transport_allowance=transport_allowance,
                            meal_allowance=meal_allowance,
                            utility_allowance=utility_allowance,
                            other_allowances=other_allowances,
                            overtime=overtime,
                            arrears=arrears,
                            staff_loan_deduction=staff_loan_deduction,
                            rent_deduction=rent_deduction,
                            medical_deduction=medical_deduction,
                            pension_contribution=pension_contribution,
                            shortage_deduction=shortage_deduction,
                            development_levy=development_levy,
                            personal_income_tax=personal_income_tax,
                            no_of_days=no_of_days,
                            actual_attendance=actual_attendance,
                            salary_date=timezone.now()
                        )
                        
                        # Calculate payslip values
                        gross_pay = (
                            salary.basic_pay +
                            salary.rent_allowance +
                            salary.transport_allowance +
                            salary.meal_allowance +
                            salary.utility_allowance +
                            salary.other_allowances +
                            salary.overtime +
                            salary.arrears
                        )
                        total_deductions = (
                            salary.staff_loan_deduction +
                            salary.rent_deduction +
                            salary.medical_deduction +
                            salary.pension_contribution +
                            salary.shortage_deduction +
                            salary.development_levy +
                            salary.personal_income_tax
                        )
                        net_pay = gross_pay - total_deductions

                        # Create payslip
                        Payslip.objects.create(
                            employee=salary.employee,
                            salary=salary,
                            date=salary.salary_date,
                            gross_pay=gross_pay,
                            total_deductions=total_deductions,
                            net_pay=net_pay
                        )
                        success_count += 1
                        
                    except (ValueError, TypeError) as e:
                        # Value conversion error (e.g. non-numeric data in numeric field)
                        errors.append(f'Row {row_num}: Data type error - {str(e)}')
                        error_count += 1
                    except Exception as e:
                        errors.append(f'Row {row_num}: Error processing record - {str(e)}')
                        error_count += 1
                
                # Report results
                if success_count > 0:
                    messages.success(request, f'Successfully imported {success_count} salary record(s).')
                
                if error_count > 0:
                    error_message = f'Failed to import {error_count} record(s). First 5 errors:'
                    for i, error in enumerate(errors[:5], 1):
                        error_message += f'<br>{i}. {error}'
                    if len(errors) > 5:
                        error_message += f'<br>...and {len(errors) - 5} more errors.'
                    messages.error(request, error_message)
                
                return redirect('core:salary_list')
                
            except openpyxl.utils.exceptions.InvalidFileException:
                messages.error(request, 'The file is not a valid Excel file. Please use the template provided by the system.')
            except Exception as e:
                messages.error(request, f'An error occurred while processing the file: {str(e)}')
        else:
            messages.error(request, 'Failed to upload payroll data. Please check the file format.')
    else:
        form = PayrollUploadForm()
    return render(request, 'core/import_payroll.html', {'form': form})

@login_required
def bulk_send_payslips(request):
    if request.method == 'GET' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        month = request.GET.get('month')
        year, month = map(int, month.split('-'))
        salaries = Salary.objects.filter(salary_date__year=year, salary_date__month=month)
        count = salaries.count()
        return JsonResponse({'count': count})

    if request.method == 'POST':
        month = request.POST.get('month')
        year, month = map(int, month.split('-'))
        salaries = Salary.objects.filter(salary_date__year=year, salary_date__month=month)
        for salary in salaries:
            payslip = Payslip.objects.get(salary=salary)
            html = generate_payslip_pdf(payslip)
            pdf = html.write_pdf()

            email = EmailMessage(
                subject=f'Payslip for {payslip.employee.name}',
                body=f"Dear {payslip.employee.name},\nPlease find attached your payslip for the month of {payslip.date.strftime('%B')}.",
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[payslip.employee.email],
            )
            email.attach(f'payslip_{payslip.employee.employee_code}_{payslip.date}.pdf', pdf, 'application/pdf')
            email.send()
        messages.success(request, f'{salaries.count()} payslips sent successfully.')
        return redirect('core:dashboard')
    return render(request, 'core/bulk_send_payslips.html')


@login_required
def monthly_payroll_summary(request):
    form = MonthlyPayrollSummaryForm(request.POST or None)
    report_data = None

    if request.method == 'POST' and form.is_valid():
        month = form.cleaned_data['month']
        staff_type = form.cleaned_data['staff_type']
        location = form.cleaned_data['location']

        payslips = Payslip.objects.filter(date__year=month.year, date__month=month.month)

        if staff_type:
            payslips = payslips.filter(employee__staff_type=staff_type)
        if location:
            payslips = payslips.filter(employee__location=location)

        if payslips.exists():
            report_data = {
                'total_salaries': payslips.aggregate(Sum('gross_pay'))['gross_pay__sum'] or 0,
                'total_deductions': payslips.aggregate(Sum('total_deductions'))['total_deductions__sum'] or 0,
                'total_net_pay': payslips.aggregate(Sum('net_pay'))['net_pay__sum'] or 0,
                'payslips': payslips
            }
        else:
            messages.info(request, "No data found for the specified month, staff type, and location.")

    return render(request, 'core/monthly_payroll_summary.html', {'form': form, 'report_data': report_data})


@login_required
def employee_compensation_report(request):
    form = EmployeeCompensationReportForm(request.POST or None)
    report_data = None

    if request.method == 'POST' and form.is_valid():
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        employees = form.cleaned_data['employees']
        staff_type = form.cleaned_data['staff_type']
        location = form.cleaned_data['location']

        payslips = Payslip.objects.filter(date__range=[start_date, end_date])

        if employees:
            payslips = payslips.filter(employee__in=employees)
        
        if staff_type:
            payslips = payslips.filter(employee__staff_type=staff_type)
        
        if location:
            payslips = payslips.filter(employee__location=location)

        if payslips.exists():
            report_data = {
                'payslips': payslips
            }
        else:
            messages.info(request, "No data found for the specified date range, employees, staff type, and location.")

    return render(request, 'core/employee_compensation_report.html', {'form': form, 'report_data': report_data})

@login_required
def help_section(request):
    return render(request, 'core/help.html')